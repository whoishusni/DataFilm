import requests as rq
import datetime as dt
from pyfiglet import Figlet
import os
import time
import platform
from openpyxl import Workbook, load_workbook

#variable film
API_KEY = ''
BASE_URL = 'https://api.themoviedb.org/3'
IMAGE_URL = 'https://image.tmdb.org/t/p/w500'

format_sekarang = dt.datetime.now()
sekarang = format_sekarang.strftime("%d - %B - %y")

def banner_show():
    banner = Figlet(font='big')
    banner_show = banner.renderText('DataFilm')
    return print(banner_show)

def write_trending_movie():
    wb = Workbook()
    ws = wb.active
    ws.append(['Film Trending Minggu Ini Tanggal {}'.format(str(sekarang))])
    ws.append(['Judul','Tanggal Rilis','Skor','Sinopsis','Image'])
    wb.save('movie_trending.xlsx')

def write_trending_tv():
    wb = Workbook()
    ws = wb.active
    ws.append(['Film Trending Minggu Ini Tanggal {}'.format(str(sekarang))])
    ws.append(['Judul','Tayang Perdana','Skor','Sinopsis','Image'])
    wb.save('series_trending.xlsx')

def write_indonesia():
    wb = Workbook()
    ws = wb.active
    ws.append(['Judul','Tanggal Rilis','Skor','Sinopsis','Image'])
    wb.save('indonesia.xlsx')


def write_search_movie(keyword):
    wb = Workbook()
    ws = wb.active
    ws.append(['Judul','Tanggal Rilis','Skor','Sinopsis','Image'])
    wb.save('movie_{}.xlsx'.format(keyword))

def write_search_tv(keyword):
    wb = Workbook()
    ws = wb.active
    ws.append(['Judul','Tayang Perdana','Skor','Sinopsis','Image'])
    wb.save('series_{}.xlsx'.format(keyword))

def terminal_clear():
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')
  
def trending_movie():
    write_trending_movie()
    terminal_clear()
    banner_show()
    TRENDING_ENDPOINT = f'/trending/movie/week?api_key={API_KEY}&language=id-ID'
    response = rq.get(BASE_URL+TRENDING_ENDPOINT).json()
    print('Daftar 10 Film Box Office yang Trending Minggu Ini {}' .format(str(sekarang)))
    for i in response['results']:
        try:
            print('Mengambil Judul: {}'.format(i['title']))
            wb = load_workbook('movie_trending.xlsx')
            ws = wb.active
            data = [i['original_title'],i['release_date'],i['vote_average'],i['overview'],IMAGE_URL+i['poster_path']]
            ws.append(data)
            wb.save('movie_trending.xlsx')
        except:
            print('Error Pengambilan Data \"{}\", sedang Melewati, harap tunggu sebentar...\n'.format(i['title']))
            #time.sleep(1)
            continue
    print('data selesai diambil\n')

def trending_tv():
    write_trending_tv()
    terminal_clear()
    banner_show()
    TRENDING_ENDPOINT = f'/trending/tv/week?api_key={API_KEY}&language=id-ID'
    response = rq.get(BASE_URL+TRENDING_ENDPOINT).json()
    print('Daftar TV Series Yang Trending Sekarang Tanggal {}\n'.format(sekarang))
    for i in response['results']:
        try:
            print('Mengambil Judul : {}'.format(i['name']))
            wb = load_workbook('series_trending.xlsx')
            ws = wb.active
            data = [i['original_name'],i['first_air_date'],i['vote_average'],i['overview'],IMAGE_URL+i['poster_path']]
            ws.append(data)
            wb.save('series_trending.xlsx')
        except:
            print('Error Pengambilan Data \"{}\", sedang Melewati, harap tunggu sebentar...\n'.format(i['name']))
            #time.sleep(1)
            continue
    print('Data Selesai Diambil\n')
      
def cari_movie(query):
    terminal_clear()
    banner_show()
    write_search_movie(query)
    SEARCH_ENDPOINT = f'/search/movie?api_key={API_KEY}&language=id-ID&query={query}&page=1&include_adult=false'
    response = rq.get(BASE_URL+SEARCH_ENDPOINT).json()
    print('Mengambil {} Judul Film Yang Berkaitan\n'.format(response['total_results']))
    for hal in range(1,response['total_pages']+1):
        SEARCH_ENDPOINT2 = f'/search/movie?api_key={API_KEY}&language=id-ID&query={query}&page={hal}&include_adult=false'
        responses = rq.get(BASE_URL+SEARCH_ENDPOINT2).json()
        for j in responses['results']:
            try:
                print('Mengambil Judul : {}'.format(j['title']))
                wb = load_workbook('movie_{}.xlsx'.format(query))
                ws = wb.active
                data = [j['original_title'],j['release_date'],j['vote_average'],j['overview'],IMAGE_URL+j['poster_path']]
                ws.append(data)    
                wb.save('movie_{}.xlsx'.format(query))
            except:
                print('Error Pengambilan Data \"{}\", sedang Melewati, harap tunggu sebentar...\n'.format(j['original_title']))
                #time.sleep(1)
                continue
    print('data selesai diambil\n')

def cari_series(query):
    terminal_clear()
    banner_show()
    write_search_tv(query)
    SEARCH_ENDPOINT = f'/search/tv?api_key={API_KEY}&language=id-ID&query={query}&page=1&include_adult=false'
    response = rq.get(BASE_URL+SEARCH_ENDPOINT).json()
    print('Mengambil {} Judul Series Yang Berkaitan\n'.format(response['total_results']))
    for hal in range(1,response['total_pages']+1):
        SEARCH_ENDPOINT2 = f'/search/tv?api_key={API_KEY}&language=id-ID&query={query}&page={hal}&include_adult=false'
        responses = rq.get(BASE_URL+SEARCH_ENDPOINT2).json()
        for j in responses['results']:
            try:
                print('Mengambil Judul : {}'.format(j['name']))
                wb = load_workbook('series_{}.xlsx'.format(query))
                ws = wb.active
                data = [j['original_name'],j['first_air_date'],j['vote_average'],j['overview'],IMAGE_URL+j['poster_path']]
                ws.append(data)   
                wb.save('series_{}.xlsx'.format(query))
            except:
                print('Error Pengambilan Data \"{}\", sedang Melewati, harap tunggu sebentar...\n'.format(j['original_name']))
                #time.sleep(1)
                continue
    print('data selesai diambil\n')

def indonesia():
    write_indonesia()
    terminal_clear()
    banner_show()
    SEARCH_ENDPOINT = f'/discover/movie?api_key={API_KEY}&include_adult=false&page=1&with_original_language=id&language=id-ID'
    response = rq.get(BASE_URL+SEARCH_ENDPOINT).json()
    print('Mengambil {} Judul Film Indonesia\n'.format(response['total_results']))
    for hal in range(1,response['total_pages']+1):
        SEARCH_ENDPOINT2 = f'/discover/movie?api_key={API_KEY}&include_adult=false&page={hal}&with_original_language=id&language=id-ID'
        responses = rq.get(BASE_URL+SEARCH_ENDPOINT2).json()
        for j in responses['results']:
            try:
                print('Mengambil Judul : {}'.format(j['original_title']))
                wb = load_workbook('indonesia.xlsx')
                ws = wb.active
                data = [j['original_title'],j['release_date'],j['vote_average'],j['overview'],IMAGE_URL+j['poster_path']]
                ws.append(data)   
                wb.save('indonesia.xlsx')
            
            except:
                print('Error Pengambilan Data \"{}\", sedang Melewati, harap tunggu sebentar...\n'.format(j['original_title']))
                #time.sleep(1)
                pass
    print('data selesai diambil\n')

def main():
    terminal_clear()
    banner_show()
    print('Created by : Husni')
    print('version    : 1.0.1\n')
    print(' Menu Utama '.center(50,'='))
    print('1. Daftar Film Trending Mingguan')
    print('2. Daftar Serial TV Trending Mingguan')
    print('3. Pencarian Film')
    print('4. Pencarian Serial TV')
    print('5. Daftar Semua Film Indonesia')
    print('6. Exit')
    print('\n')
    menu = input('Pilih Menu [1-6]: ')
    if menu == '1':
        trending_movie()
    elif menu == '2':
        trending_tv()
    elif menu == '3':
        query = input('Masukkan Judul Film yang mau dicari ? ')
        cari_movie(query)
    elif menu == '4':
        query = input('Masukkan Judul Series yang mau dicari ? ')
        cari_series(query)
    elif menu == '5':
        indonesia()
    elif menu == '6':
        print('Exit....')
        time.sleep(0.5)
        os._exit(os.EX_OK)
    else:
        print('Menu Nomor {} Tidak Ditemukan\nSilahkan Coba Lagi'.format(menu))
        time.sleep(2)
        main()

if __name__ == '__main__':
    main()